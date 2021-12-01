import pandas as pd
from openpyxl import load_workbook
def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames
def find_street(db,adresse,street):
  fl=[]
  for i in db[adresse]:
    s=str(i).lower()
    if s.find(street.lower()) != -1:
      fl.append(True)
    elif s.find(street.lower()) == -1:
      fl.append(False)
  return fl
def count_files():
    import pathlib
    initial_count = 0
    for path in pathlib.Path(".").iterdir():
        if path.is_file():
            initial_count += 1
    return initial_count
df = pd.DataFrame([], columns=["NPN","Adresse","CP","Ville","Num","T-age"])
nb=count_files()
for k in range(1,nb):
  sheets=get_sheetnames_xlsx(str(k)+".xlsx")
  for i in sheets:
    d=pd.read_excel(str(k)+".xlsx",header=None,sheet_name=i)
    df=pd.concat([df, d], ignore_index = True, axis = 0)
    df=df[df.columns[0:6]]
df.columns=["NPN","Adresse","CP","Ville","Num","T-age"]
df.to_excel("idf.xlsx")
streets=[str(i) for i in pd.read_excel("poors_filter_idf.xlsx").R_AV_M]
db=pd.read_excel("idf.xlsx")
df = pd.DataFrame([], columns=db.columns)
for k in streets:
  d=db[find_street(db,"Adresse",k)]
  df=pd.concat([df, d], ignore_index = True, axis = 0)
del df['Unnamed: 0']
df.to_excel("poor3.xlsx")